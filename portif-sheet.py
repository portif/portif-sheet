#!/usr/bin/env python3

from argparse import ArgumentParser
from csv import DictReader

from openpyxl import Workbook

parser = ArgumentParser(description='Gerador de planilha para portifólio de aprendizagem.')
parser.add_argument('nome', help='Nome desejado para o portifólio.')
parser.add_argument('arquivo', help='Arquivo CSV com os campos "Matrícula" e "Nome".')
parser.add_argument('colunas', help='Lista de nomes das colunas, separados por vírgula (","), e cercada por aspas.')

args = parser.parse_args()

colunas = args.colunas.split(',')

leitor = DictReader(open(args.arquivo))

nome_matricula = {linha["Matrícula"]: linha["Nome"] for linha in leitor}


# Colunas
# colunas = ["Nome", "Assinatura", "Descrição", "Exemplo(s)", "Referências"]

# Criar um novo arquivo Excel
wb = Workbook()

# Adicionar a primeira planilha "Principal" com a lista de estudantes
principal_sheet = wb.active
principal_sheet.title = "Principal"

cabecalhos = ["Matrícula", "Nome"]
for j,cabecalho in enumerate(cabecalhos, start=1):
    principal_sheet.cell(row=1, column=j, value=cabecalho)

# Adicionar os nomes dos estudantes na planilha "Principal"
for i, estudante in enumerate(nome_matricula.items(), start=2):
    principal_sheet.cell(row=i, column=1, value=estudante[0])
    principal_sheet.cell(row=i, column=2, value=estudante[1])

# Criar as planilhas individuais para cada estudante
for matricula,nome in nome_matricula.items():
    prim_nome = nome.split()[0]
    nome_planilha = f'{matricula}-{prim_nome}'
    ws = wb.create_sheet(title=nome_planilha)
    for col_num, col_name in enumerate(colunas, start=1):
        ws.cell(row=1, column=col_num, value=col_name)

# # Modificar a planilha "Principal" para adicionar links para as planilhas correspondentes
# for i, estudante in enumerate(estudantes, start=1):
#     # Criar o link para a planilha do estudante correspondente
#     link = f"'{estudante}'!A1"
#     principal_sheet.cell(row=i, column=1).value = f'=HYPERLINK("{link}", "{estudante}")'

# Salvar o arquivo Excel
# TODO: Nomear o arquivo de saída de acordo com o nome do arquivo passado como argumento.
output_path = f'./portif-{args.nome}.xlsx'
wb.save(output_path)


